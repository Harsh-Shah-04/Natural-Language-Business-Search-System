"""
M1.2 raw ingestion + M1.3 embedding backfill.

Parses the provided .xlsx dataset, inserts all 14 fields per business into
MongoDB, then computes a BAAI/bge-small-en-v1.5 embedding per business from
the fields with semantic signal (see app/embeddings.py) and backfills the
384-dim `embedding` field before insert.

Safe to re-run: clears the businesses collection and reinserts fresh each
time, so repeated runs always converge on exactly one document per row of
the source file.

Requires MONGODB_URI in .env (see backend/.env.example).

Run: uv run python scripts/seed.py [path/to/dataset.xlsx]
"""

import sys
import time
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from app.db import get_db
from app.embeddings import build_embedding_text, embed_texts

# Repo layout: backend/scripts/seed.py -> backend/ -> repo root -> dataset.
DEFAULT_DATASET_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "Business_Matchmaking_Test_Dataset_V2_120_Companies.xlsx"
)

# Source column order (as it appears in the .xlsx header row) mapped to the
# snake_case document field names established in M1.1's README.
COLUMNS = [
    "business_name",
    "nature",
    "industry",
    "sub_category",
    "city",
    "state",
    "contact_person",
    "email",
    "website",
    "phone",
    "business_description",
    "products_services",
    "keywords",
    "specialties",
]


def parse_rows(xlsx_path: Path) -> list[dict]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    if len(header) != len(COLUMNS):
        raise ValueError(
            f"Expected {len(COLUMNS)} columns in {xlsx_path.name}, found {len(header)}: {header}"
        )

    documents = []
    for row in rows:
        if all(cell is None for cell in row):
            continue  # skip fully blank rows
        documents.append(dict(zip(COLUMNS, row)))
    return documents


def main() -> int:
    load_dotenv()

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_DATASET_PATH
    if not xlsx_path.exists():
        print(f"FAIL: dataset not found at {xlsx_path}")
        return 1

    documents = parse_rows(xlsx_path)

    names = [doc["business_name"] for doc in documents]
    if any(not name for name in names):
        print("FAIL: at least one row is missing business_name")
        return 1
    if len(names) != len(set(names)):
        print("FAIL: duplicate business_name values found in the source file")
        return 1

    embed_start = time.perf_counter()
    texts = [build_embedding_text(doc) for doc in documents]
    vectors = embed_texts(texts)
    embed_elapsed = time.perf_counter() - embed_start

    for doc, vector in zip(documents, vectors):
        doc["embedding"] = vector

    docs_per_sec = len(documents) / embed_elapsed if embed_elapsed > 0 else float("inf")
    print(
        f"Embedded {len(documents)} businesses in {embed_elapsed:.2f}s "
        f"({docs_per_sec:.1f} docs/sec, model load included)"
    )

    db = get_db()
    businesses = db["businesses"]

    businesses.create_index("business_name", unique=True)
    businesses.delete_many({})
    try:
        businesses.insert_many(documents)
    except Exception as e:
        print(
            f"FAIL: insert_many raised after clearing the collection — "
            f"businesses now holds {businesses.count_documents({})} of {len(documents)} "
            f"documents. Re-run this script to retry from a clean state. "
            f"Underlying error: {e}"
        )
        return 1

    count = businesses.count_documents({})
    if count != len(documents):
        print(f"FAIL: inserted {count} documents, expected {len(documents)}")
        return 1

    print(f"PASS: seeded {count} businesses from {xlsx_path.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
